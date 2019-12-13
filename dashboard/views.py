from django.shortcuts import render, redirect, get_object_or_404 
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.models import User
from django.shortcuts import render, HttpResponse
from dashboard.models import *
import requests
import json
import openpyxl
import re

@login_required()
def dashboard(request):
    if 'dic' in request.session:#-----------------------------------------
        del request.session['dic']
    if 'msg' in request.session:
        del request.session['msg']
    usercount = User.objects.all().count()
    badcount=0;goodcount =0
    months=[]
    if request.user.is_superuser:
        count = transactions.objects.all().count()
        for i in range(12):
            months.append(transactions.objects.filter(created_at__month=(i+1)).count())
        for trans in transactions.objects.all():
            if trans.pwned=="Bad!":
                badcount+=1
            else:
                goodcount+=1
    else:
        count = transactions.objects.filter(user_id = request.user).count()
        for i in range(12):
            months.append(transactions.objects.filter(created_at__month=(i+1),user_id=request.user).count())
        for trans in transactions.objects.filter(user_id=request.user):
            if trans.pwned=="Bad!":
                badcount+=1
            else:
                goodcount+=1
    piedata = [badcount, goodcount]
    areadata = months
    return render(request,'dashboard/index.html',{"count": count ,"usercount": usercount,"piedata":piedata,"areadata":areadata})

@login_required()
def check(request):
    if 'dic' in request.session:
        return render(request,'dashboard/check.html',request.session['dic'])
    if 'msg' in request.session:
        del request.session['msg']
    else:
        return render(request,'dashboard/check.html')


@login_required()
def records(request):
    if 'dic' in request.session:
        del request.session['dic']
    if User.objects.get(id= request.user.id).is_superuser :
        return render(request,'dashboard/records.html',{'myrecords':transactions.objects.all() })
    else:
        return render(request,'dashboard/records.html',{'myrecords':transactions.objects.filter(user_id=request.user) })

def deleterecord(request,id):
        del_record = transactions.objects.get(id=id)
        del_record.delete()
        messages.success(request,"User has been successfully deleted")
        request.session['msg']=  "<div class='alert alert-danger' role='alert'>Record  deleted Successfully !</div>"
        return redirect('records')
    
@login_required()
def userslist(request):
    if 'dic' in request.session:
        del request.session['dic']
    if 'msg' in request.session:
        del request.session['msg']
    return render(request,'dashboard/userslist.html',{'userslist':User.objects.all() })


def deleteuser(request,id):
    del_user = get_object_or_404(User, pk=id)
    del_user.delete()
    messages.info(request,'Record has been successfully deleted ')
    return redirect('userslist')
     
# def edit_user_profile(request):
#     if request.method == "POST":
#         userID = request.POST['userId']
#         user_to_update = User.objects.get(id=int(userID))
#         f = request.POST['firstName']
#         last = request.POST['lastName']
#         eml = request.POST['e-mail']
#         user_to_update.first_name = f
#         user_to_update.last_name = last
#         user_to_update.email = eml
#         user_to_update.save()
#         print(user_to_update.first_name)
#         return render(request,'dashboard/userslist.html')
#     else:
#         return render(request,'dashboard/userslist.html')

@login_required()
def edit_profile(request):
    if request.method == "POST":
        user_username = request.POST['userId']
        userID = User.objects.get(username=user_username).pk
        user_to_update = User.objects.get(id=userID)
        f = request.POST['FirstName']
        last = request.POST['LastName']
        eml = request.POST['Email']
        user_to_update.first_name = f
        user_to_update.last_name = last
        user_to_update.email = eml
        user_to_update.save()
        return render(request,'dashboard/settings.html',{'msg':"<div class='alert alert-success' role='alert'>You're successfully update your file!</div>"})
    else:
        if 'dic' in request.session:
            del request.session['dic']
        if 'msg' in request.session:
            del request.session['msg']
        return render(request,'dashboard/settings.html')


        
def one_email(request):#----------------------------------
    #when enter emails in txt form 
    if request.method == "POST":
        email = request.POST['email']
        emails = email.replace(" ", "").split(';')
        emails_list = list()
        regex = '^\w+([\.-]?\w+)*@\w+([\.-]?\w+)*(\.\w{2,3})+$'
        flag = False
        for email in emails:
            if (re.search(regex,email)):
                dec = result(request,email)
                dec['email'] = email
                emails_list.append(dec)
                if dec['pwned'] == "Bad!":
                    flag = True
                transactions.objects.create(email=dec['email'],pwned=dec['pwned'],breaches=dec['breaches'],user=request.user)
            else:
                request.session['dic']=  {"msg1":"<div class='alert alert-danger' role='alert'>Please, Enter vaild Email syntax e.g(example@example.com)</div>"}
                return redirect('check')
        request.session['dic']= {"emails_list":emails_list, "flag": flag}
        return redirect("check" )
    else:
        request.session['dic']= {}
        return redirect("check")

def excel_email(request):#-------------------------------
    if "GET" == request.method:
        request.session['dic']={}
        return redirect('check')
    else:
        # exciption for file extension
        try:
            excel_file = request.FILES["excel_file"]
            wb = openpyxl.load_workbook(excel_file)
            sheets = wb.sheetnames
            emails_list = list()
            # for validating an Email
            regex = '^\w+([\.-]?\w+)*@\w+([\.-]?\w+)*(\.\w{2,3})+$'
            flag = False
            for sheet in sheets:
                worksheet = wb[sheet]
                # iterating over the rows 
                for row in worksheet.iter_rows():
                    for cell in row:
                        if (re.search(regex,str(cell.value))):
                            dec = result(request,cell.value)
                            dec['email'] = str(cell.value)
                            emails_list.append(dec)
                            if dec['pwned'] == "Bad!":
                                flag = True
                            # emails_list.append(str(cell.value))
                            transactions.objects.create(email=dec['email'],pwned=dec['pwned'],breaches=dec['breaches'],user=request.user)
            request.session['dic']={"emails_list":emails_list, "flag": flag}
            return redirect('check')
        except :
            request.session['dic']={"msg2":"<div class='alert alert-danger' role='alert'>Please, upload xlsx/xlsm/xltx/xltm files!</div>"}
            return redirect('check')
            
def result(request,email):   
    #Get API function   
    url = "https://haveibeenpwned.com/api/v3/breachedaccount/"+email
    headers = {
        'hibp-api-key': "5c1b56d410604568a61adf4295121551"
        }
    try:
        rs = requests.request("GET", url, headers=headers).json()
    except json.decoder.JSONDecodeError as e:
        print(e)
        rs = []
    if len(rs) == 0:
        pwned = "Good!"
    else:
        pwned = "Bad!"
    
    response = {'pwned': pwned, 'breaches': rs}
    return response
